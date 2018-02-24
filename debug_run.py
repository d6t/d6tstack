"""Debug file.

    Note:
        You can pretty much ignore this


"""

import glob
import combiner.combine_files
import combiner.helpers
from importlib import reload
from combiner.test_combine import create_files_df_clean, create_files_df_clean_combine
import csv


fname_list = glob.glob('combiner/test-data/input/test-data-input-csv-colmismatch2-*.csv')
r = combiner.combine_files.combine_files(fname_list,'',cfg_return_df=True)
print(r['status'])
print(r['columns_files'])


fname_col = {'combiner/test-data/input/test-data-input-csv-colmismatch2-mar.csv': ['date', 'sales', 'cost', 'profit', 'profit0', 'profit1', 'profit2', 'profit3', 'profit4', 'profit5', 'profit6', 'profit7', 'profit8', 'profit9', 'profit10', 'profit11', 'profit12', 'profit13', 'profit14', 'file_name'], 'combiner/test-data/input/test-data-input-csv-colmismatch2-feb.csv': ['date', 'sales', 'cost', 'profit', 'file_name'], 'combiner/test-data/input/test-data-input-csv-colmismatch2-jan.csv': ['date', 'sales', 'cost', 'profit', 'file_name']}
import combiner.combine_files
reload(combiner.combine_files)
combiner.combine_files.column_mismatch_dict(fname_col)

#print(combine_files(fname_list,'.'))
'''
fname_list = glob.glob('test-data/test-data-input-xls-sing-*.xls')

r = combine_files(fname_list,'test-data/test.xls',cfg_return_df=True)
df = r['data']
df = df.sort_values('date').drop(['file_name'], axis=1)

df_chk = create_files_df_clean_combine()

a =1
'''

fname_list = glob.glob('test-data/input/test-data-input-csv-*.csv')
fname_list = glob.glob('test-data/input/test-data-input-noheader-csv-*.csv')

import openpyxl

fname = '/mnt/data/DF_USTotal_Sedol_Rank_20170501_20170531.xlsx'
fname = '/mnt/data/test-data-input-xls-sing-mar.xlsx'
fh = openpyxl.load_workbook(fname, read_only=True)
fh.sheetnames
ws = fh[fh.sheetnames[0]]
pd.DataFrame(ws.values)

irow = 0
for row in ws.rows:
    print([c.value for c in row])
    irow += 1
    if irow>3:
        break

df = DataFrame(ws.values)

import pandas as pd
df=pd.read_excel(fname,nrows=10)
df.head()

